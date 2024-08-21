# -*- coding: utf-8 -*-
"""
Created on Tue Jul 16 11:33:10 2024

@author: klecznar
"""

import sys, time
import openpyxl
import logging
import pandas as pd
from datetime import datetime

# lists to hold data from Excel
# portfolio
ref_list_port = []
date_list_port = []
date_list_port1 = []  # list to be converted to df to remove duplicates

# extract
ref_list_ext = []
date_list_ext = []
date_list_ext1 = []  # list to be converted to df to remove duplicates
unique_ext_pn = []
unique_ext_date = []

# tracker
list_tr1 = []  # list with duplicates
pn_conc = []
date_conc = []

# Quality ePO
item_list_ePO = []
vendor_list_ePO = []



def launch_portfolio():
    
    """
    Function analyses Portfolio file and returns lists with PNs and target dates
    """
    
    try:
        excel = openpyxl.load_workbook('PortofolioOrdersDataGridDetail.xlsx')
        sheet1 = excel["PortofolioOrdersDataGridDetail"]
        
        for i in range(1048575):  # 1,048,576 is max number of rows in excel
            ref_item_port = sheet1.cell(row=2 + i, column=1).value
            date_item_port = sheet1.cell(row=2 + i, column=5).value 
            if ref_item_port is None or ref_item_port == '':
                break
            if ref_item_port is not None and ref_item_port != '':
                # get all PNs with duplicates
                ref_list_port.append(ref_item_port)
                date_item_port_con = str(date_item_port).rstrip(' 00:00:00')
                date_item_port_con2 = datetime.strptime(date_item_port_con, '%Y-%m-%d').strftime('%d/%m/%Y')
                date_list_port1.append(date_item_port_con2)

                        
    except Exception:
        logging.exception("Could not open the Portfolio file! Please amend filename to PortofolioOrdersDataGridDetail")
        time.sleep(10)
        sys.exit()
        
    return ref_list_port, date_list_port1
        

def remove_duplicates_port():
    
    """
    Function removes duplicates from list with PNs and keeps the list order with dates
    """
    
    global ref_list_port, date_list_port1, date_list_port
    
    # convert data type to str
    date_list_port = map(str, date_list_port1)
        
    # convert to DataFrame
    df = pd.DataFrame(list(zip(ref_list_port, date_list_port)),
                   columns =['PN', 'Date'])

    # drop duplicates
    df = df.sort_values('Date').drop_duplicates('PN', keep='last')

    # convert df to lists
    ref_list_port = df['PN'].values.tolist()
    date_list_port = df['Date'].values.tolist()

    
    return ref_list_port, date_list_port
        
        
def launch_extract():
    
    """
    Function analyses GEAC extract file and returns lists with PNs and target dates
    """
    
    try:
        excel = openpyxl.load_workbook('DVI inf.xlsx')
        sheet2 = excel['Sheet1']
        
        for i in range(1048575):  # 1,048,576 is max number of rows in excel
            ref_item_ext = sheet2.cell(row=2 + i, column=2).value
            date_item_ext = sheet2.cell(row=2 + i, column=26).value  
            if ref_item_ext is None or ref_item_ext == '':
                break
            if ref_item_ext is not None and ref_item_ext != '':
                ref_list_ext.append(ref_item_ext)
                try:
                    # convert date format 1RRMMDD -> DD/MM/RRRR
                    date_item_ext_con = str(date_item_ext).lstrip('1')
                    date_item_ext_conv = datetime.strptime(date_item_ext_con, '%y%m%d').strftime('%d/%m/%Y')
                    date_list_ext1.append(date_item_ext_conv)
                except:
                    # if else append as is
                    date_list_ext1.append(date_item_ext)

                        
    except Exception:
        logging.exception("Could not open the GEAC extract file! Please amend filename to DVI inf")
        time.sleep(10)
        sys.exit()
    
        
    return ref_list_ext, date_list_ext1
    
    
def remove_duplicates_ext():
    
    """
    Function removes duplicates from list with PNs and keeps the list order with dates
    """
    
    global ref_list_ext, date_list_ext1, date_list_ext
    
    # convert data type to str
    date_list_ext = map(str, date_list_ext1)

    # convert to DataFrame
    df = pd.DataFrame(list(zip(ref_list_ext, date_list_ext)),
                   columns =['PN', 'Date'])

    # drop duplicates
    df = df.sort_values('Date').drop_duplicates('PN', keep='last')

    # convert df to lists
    ref_list_ext = df['PN'].values.tolist()
    date_list_ext = df['Date'].values.tolist()
    
    
    return ref_list_ext, date_list_ext



def compare_port_ext():
    
    """
    Function compares lists from the Portfolio and GEAC extract files 
    and keeps only the unique PNs
    """
    
    global ref_list_port, ref_list_ext, date_list_ext
    
    
    for item, pn, date in zip(ref_list_port, ref_list_ext, date_list_ext):
        if pn not in ref_list_port:
            unique_ext_pn.append(pn)
            unique_ext_date.append(date)
        
    return unique_ext_pn, unique_ext_date
    


def launch_tracker():
    
    """
    Function analyses Safran DVI Tracker file and updates it
    """
    
    global ref_list_port, date_list_port, unique_ext_pn, unique_ext_date, date_list_ext
    
    # GET ALL PNs FROM TRACKER
    try:
        excel = openpyxl.load_workbook("Safran DVI tracker.xlsx")
        sheet4 = excel["tracker"]
        row_count = sheet4.max_row
        
        # get all PNs from tracker
        for j in range(2, row_count):
            pn_item = sheet4.cell(row= j, column=2).value
            if pn_item is None or pn_item == '':
                break
            if pn_item is not None and pn_item != '':
                # get all PNs with duplicates
                list_tr1.append(pn_item)
                
        # remove duplicates
        list_tr = list(set(list_tr1))
        
        
        # compare tracker with portfolio
        for pn, date in zip(ref_list_port, date_list_port):
            # loop through tracker
            for j in range(2, row_count):
                pn_item = sheet4.cell(row= j, column=2).value
                if pn == pn_item:
                    date_item = str(sheet4.cell(row = j, column = 4).value).rstrip(' 00:00:00') 
                    if (date_item == "None" or date_item == '') or date_item < date:
                        sheet4.cell(column=4, row=j, value=date)       
            if pn not in list_tr:
                pn_conc.append(pn)
                date_conc.append(date)

        
        # save changes to worksheet
        excel.save("Safran DVI tracker.xlsx")
                    
        
        # compare tracker with geac extract
        for pn, date in zip(unique_ext_pn, unique_ext_date):
            # loop through tracker
            for j in range(2, row_count):
                pn_item = sheet4.cell(row= j, column=2).value
                if pn == pn_item:
                    date_item = str(sheet4.cell(row = j, column = 4).value).rstrip(' 00:00:00') 
                    if (date_item == "None" or date_item == '') or date_item < date:
                        sheet4.cell(column=4, row=j, value=date)       
            if pn not in list_tr:
                pn_conc.append(pn)
                date_conc.append(date)


    except Exception:
        logging.exception("Could not open the DVI Tracker! Please amend filename to Safran DVI tracker")
        time.sleep(10)
        sys.exit()

    # save changes to workbook
    excel.save("Safran DVI tracker.xlsx")
    
    return pn_conc, date_conc


def launch_ePO():
    
    """
    Function analyses Quality ePO file and returns lists with PNs and vendors
    """    
    
    try:
        excel = openpyxl.load_workbook('Quality ePO.xlsx', data_only=True)
        sheet3 = excel['Analysis']
        row_count = sheet3.max_row
             
        for pn in pn_conc:
            for j in range(2, row_count):
                item_ePO = sheet3.cell(row=2 + j, column=4).value
                vendor_item = sheet3.cell(row = 2 + j, column = 18).value
                if item_ePO is None or item_ePO == '':
                    break
                if item_ePO is not None and item_ePO != '':
                    if pn == item_ePO:
                        item_list_ePO.append(item_ePO)
                        vendor_list_ePO.append(vendor_item)


                        
    except Exception:
        logging.exception("Could not open the Quality ePO file! Please amend filename to Quality ePO")
        time.sleep(10)
        sys.exit()
    
        
    return item_list_ePO, vendor_list_ePO



def new_data():
    
    """
    Function creates a new Excel file with returned data
    """
      
    # create df to append to excel
    df = pd.DataFrame(
        {
            'Customer' : 'Safran Aircraft Engines',
            'Customer pn' : pn_conc,
            'Start Date' : None,
            'Target Date' : date_conc,
            'Drawing no' : None,
            'Supplier DVI no' : None,
            'Incora DVI no' : None,
            'Customer DVI no' : None,
            'Reason' : None,
            'MOD/ Part revision' : None,
            'Drawing revision' : None,
            'Class' : None,
            'Flag control' : None,
            'Type' : None,
            'DVI required' : None,
            'Required date' : None,
            'Days to date' : None,
            'Requester' : None,
            'Submitted documentation' : None,
            'Requirement' : None,
            'Special requirement' : None,
            'SAE confirmation' : None,
            'Documentation Type' : None,
            'Supplier name': None
         }
    )
      
    # append df to excel
    with pd.ExcelWriter("Safran DVI tracker2.xlsx") as writer:
        df.to_excel(writer, sheet_name="tracker", index=False)
        


def paste_vendor(): 
    
    """
    Function pastes data from the vendor list if PN matches in excel
    """
    
    global item_list_ePO, vendor_list_ePO
       
    excel = openpyxl.load_workbook('Safran DVI tracker2.xlsx')
    sheet5 = excel['tracker']
    row_count = sheet5.max_row
   
    # loop through ePO duplicates
    for pn, vendor in zip(item_list_ePO, vendor_list_ePO):
        # loop through tracker2
        for j in range(2, row_count):
            pn_item = sheet5.cell(row= j, column=2).value
            if pn == pn_item:
                sheet5.cell(column=24, row=j, value=vendor) 
    
    
    # save changes to workbook
    excel.save("Safran DVI tracker2.xlsx")
     
    
    



        
        
        
        
        
        
        
        
        
        


